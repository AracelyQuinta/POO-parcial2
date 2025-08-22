
"""- Guarda productos en Excel sin duplicarlos.
- Carga productos desde Excel con validación por fila.
- Lee y escribe archivos de texto plano con manejo de errores.
"""


from openpyxl import Workbook, load_workbook
from producto import Producto
import os

class Archivo:
    def __init__(self, productos=None):
        # Inicializa la lista de productos. Si se recibe una lista, la usa; si no, crea una vacía.
        self.productos = productos if productos else []

    def guardar_en_excel_xlsx(self, nombre_archivo):
        # Verifica si hay productos en memoria antes de guardar
        if not self.productos:
            print("No hay productos en memoria para guardar.")
            return
        try:
            if os.path.exists(nombre_archivo):
                # Si el archivo existe, lo carga
                wb = load_workbook(nombre_archivo)
                ws = wb.active

                # Elimina todas las filas de datos (desde la fila 2 en adelante), conservando el encabezado
                ws.delete_rows(2, ws.max_row)
            else:
                # Si el archivo no existe, lo crea con encabezado
                wb = Workbook()
                ws = wb.active
                ws.title = "Inventario"
                ws.append(["ID", "Nombre", "Cantidad", "Precio"])

            # Escribe cada producto en una nueva fila
            for producto in self.productos:
                ws.append([
                    producto.getid(),
                    producto.getnombre(),
                    producto.getcantidad(),
                    producto.getprecio()
                ])

            # Guarda el archivo actualizado
            wb.save(nombre_archivo)
            print(f"Se ha guardado exitosamente en '{nombre_archivo}' sin duplicados.")
        except PermissionError:
            print(f"Error: No tienes permisos para modificar '{nombre_archivo}'.")
        except Exception as e:
            print(f"Error inesperado: {type(e).__name__} - {e}")

    def cargar_desde_excel(self, nombre_archivo, hoja):
        # Limpia la lista de productos antes de cargar
        self.productos = []

        # Verifica que el archivo exista
        if os.path.exists(nombre_archivo):
            wb = load_workbook(nombre_archivo)

            # Verifica que la hoja especificada exista
            if hoja not in wb.sheetnames:
                print(f"La hoja '{hoja}' no existe.")
                return

            ws = wb[hoja]

            # Recorre las filas del archivo (ignorando el encabezado)
            for fila in range(2, ws.max_row + 1):
                try:
                    # Extrae los datos de cada celda
                    id_ = str(ws.cell(row=fila, column=1).value)
                    nombre = str(ws.cell(row=fila, column=2).value)
                    cantidad = int(ws.cell(row=fila, column=3).value)
                    precio = float(ws.cell(row=fila, column=4).value)

                    # Crea un objeto Producto y lo añade a la lista
                    producto = Producto(id_, nombre, cantidad, precio)
                    self.productos.append(producto)
                except Exception as e:
                    # Si hay error en una fila, lo reporta pero continúa con las demás
                    print(f"Error al leer fila {fila}: {e}")

    def leer_archivo(ruta):
        # Lee el contenido de un archivo de texto plano
        try:
            with open(ruta, 'r', encoding='utf-8') as archivo:
                contenido = archivo.read()
                print("Archivo leído correctamente.")
                return contenido
        except FileNotFoundError:
            print(f"Error: El archivo '{ruta}' no existe.")
        except PermissionError:
            print(f"Error: No tienes permisos para acceder al archivo '{ruta}'.")
        except Exception as e:
            print(f"Error inesperado: {type(e).__name__} - {e}")

    def escribir_archivo(ruta, datos):
        # Escribe datos en un archivo de texto plano
        try:
            with open(ruta, 'w', encoding='utf-8') as archivo:
                archivo.write(datos)
                print("Datos escritos correctamente.")
        except PermissionError:
            print(f"Error: No tienes permisos para escribir en '{ruta}'.")
        except Exception as e:
            print(f"Error inesperado: {type(e).__name__} - {e}")
