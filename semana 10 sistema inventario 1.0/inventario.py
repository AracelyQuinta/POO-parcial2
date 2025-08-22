from openpyxl import load_workbook
from producto import Producto
from manipulacion_de_archivo import Archivo

class Inventario:
    def __init__(self):
        # Lista en memoria para almacenar productos temporalmente
        self.productos = []

    def añadir_producto(self):
        # Solicita ID del nuevo producto
        nuevo_id = input("Ingrese el ID: ").strip()

        # Verifica si el ID ya existe en la lista en memoria
        if any(p.getid() == nuevo_id for p in self.productos):
            print(f"Ya existe un producto con el ID {nuevo_id}.")
            return

        # Solicita nombre, cantidad y precio
        nombre = input("Nombre del producto: ").strip()
        try:
            cantidad = int(input("Cantidad del producto: "))
            precio = float(input("Precio del producto: "))
        except ValueError:
            print("Cantidad o precio inválido. Producto no añadido.")
            return

        # Crea el producto y lo añade a la lista
        nuevo_producto = Producto(nuevo_id, nombre, cantidad, precio)
        self.productos.append(nuevo_producto)
        print(f"Producto '{nombre}' añadido correctamente.")

    def mostrar_productos_guardados(self):
        # Muestra los productos que están en memoria
        if self.productos:
            for producto in self.productos:
                print(producto)
        else:
            print("No hay productos guardados en memoria.")

    def eliminar_producto_en_excel(self, nombre_archivo, hoja, id_a_eliminar):
        try:
            # Carga el archivo Excel
            wb = load_workbook(nombre_archivo)
            if hoja not in wb.sheetnames:
                print(f"La hoja '{hoja}' no existe en el archivo.")
                return

            ws = wb[hoja]
            fila_a_eliminar = None

            # Busca la fila que contiene el ID a eliminar
            for fila in range(2, ws.max_row + 1):
                id_en_excel = str(ws.cell(row=fila, column=1).value)
                if id_en_excel == id_a_eliminar:
                    fila_a_eliminar = fila
                    break

            # Elimina la fila si se encontró
            if fila_a_eliminar:
                ws.delete_rows(fila_a_eliminar)
                wb.save(nombre_archivo)
                print(f"Producto con ID {id_a_eliminar} eliminado del archivo Excel.")
            else:
                print(f"No se encontró el ID {id_a_eliminar} en el archivo Excel.")
        except PermissionError:
            print(f"No tienes permisos para modificar '{nombre_archivo}'.")
        except Exception as e:
            print(f"Error inesperado al eliminar: {type(e).__name__} - {e}")

    def actualizar_producto_por_id(self, nombre_archivo, hoja, id_a_actualizar):
        try:
            # Carga el archivo Excel
            wb = load_workbook(nombre_archivo)
            if hoja not in wb.sheetnames:
                print(f"La hoja '{hoja}' no existe.")
                return

            ws = wb[hoja]
            fila_a_actualizar = None

            # Busca la fila que contiene el ID a actualizar
            for fila in range(2, ws.max_row + 1):
                id_en_excel = str(ws.cell(row=fila, column=1).value)
                if id_en_excel == id_a_actualizar:
                    fila_a_actualizar = fila
                    break

            if fila_a_actualizar is None:
                print(f"No se encontró el ID {id_a_actualizar}.")
                return

            # Muestra los datos actuales del producto
            print(f"\nProducto con ID {id_a_actualizar} encontrado en la fila {fila_a_actualizar}.")
            nombre_actual = ws.cell(row=fila_a_actualizar, column=2).value
            cantidad_actual = ws.cell(row=fila_a_actualizar, column=3).value
            precio_actual = ws.cell(row=fila_a_actualizar, column=4).value

            print(f"Nombre actual: {nombre_actual}")
            print(f"Cantidad actual: {cantidad_actual}")
            print(f"Precio actual: {precio_actual}")

            # Opciones de actualización
            print("\n¿Qué desea actualizar?")
            print("1. Nombre")
            print("2. Cantidad")
            print("3. Precio")
            print("4. Todo")
            print("5. Cancelar")

            entrada = input("Ingrese las opciones separadas por coma (ej. 1,3 o 4): ").strip()
            opciones = [op.strip() for op in entrada.split(",")]

            if "5" in opciones:
                print("Cambios cancelados. No se guardó nada.")
                return

            if "4" in opciones:
                opciones = ["1", "2", "3"]

            # Actualiza nombre si se seleccionó
            if "1" in opciones:
                nuevo_nombre = input("Ingrese el nuevo nombre: ").strip()
                if nuevo_nombre:
                    ws.cell(row=fila_a_actualizar, column=2).value = nuevo_nombre
                    print("Nombre actualizado.")
                else:
                    print("Nombre vacío. No se actualizó.")

            # Actualiza cantidad si se seleccionó
            if "2" in opciones:
                try:
                    nueva_cantidad = int(input("Ingrese la nueva cantidad: ").strip())
                    ws.cell(row=fila_a_actualizar, column=3).value = nueva_cantidad
                    print("Cantidad actualizada.")
                except ValueError:
                    print("Cantidad inválida. No se actualizó.")

            # Actualiza precio si se seleccionó
            if "3" in opciones:
                try:
                    nuevo_precio = float(input("Ingrese el nuevo precio: ").strip())
                    ws.cell(row=fila_a_actualizar, column=4).value = nuevo_precio
                    print("Precio actualizado.")
                except ValueError:
                    print("Precio inválido. No se actualizó.")

            # Guarda los cambios en el archivo
            wb.save(nombre_archivo)
            print("Cambios guardados correctamente.")
        except PermissionError:
            print(f"No tienes permisos para modificar '{nombre_archivo}'.")
        except Exception as e:
            print(f"Error inesperado al actualizar: {type(e).__name__} - {e}")

    def buscar_productos_por_nombre(self, nombre_archivo, hoja, texto_busqueda):
        try:
            # Carga el archivo Excel
            wb = load_workbook(nombre_archivo)
            if hoja not in wb.sheetnames:
                print(f"La hoja '{hoja}' no existe.")
                return

            ws = wb[hoja]
            texto_busqueda = texto_busqueda.lower().strip()
            productos_encontrados = []

            # Busca coincidencias parciales en el nombre
            for fila in range(2, ws.max_row + 1):
                nombre = ws.cell(row=fila, column=2).value
                if nombre and texto_busqueda in nombre.lower():
                    producto = {
                        "ID": ws.cell(row=fila, column=1).value,
                        "Nombre": nombre,
                        "Cantidad": ws.cell(row=fila, column=3).value,
                        "Precio": ws.cell(row=fila, column=4).value
                    }
                    productos_encontrados.append(producto)

            # Muestra los resultados encontrados
            if productos_encontrados:
                print(f"\nSe encontraron {len(productos_encontrados)} producto(s) que coinciden con '{texto_busqueda}':\n")
                for p in productos_encontrados:
                    print(f"| ID: {p['ID']} | Nombre: {p['Nombre']} | Cantidad: {p['Cantidad']} | Precio: ${p['Precio']:.2f} |")
            else:
                print(f"\nNo se encontraron productos que coincidan con '{texto_busqueda}'.")
        except Exception as e:
            print(f"Error al buscar productos: {type(e).__name__} - {e}")

    def mostrar_todos_los_productos(self, nombre_archivo, hoja):
        try:
            # Carga el archivo Excel
            wb = load_workbook(nombre_archivo)
            if hoja not in wb.sheetnames:
                print(f"La hoja '{hoja}' no existe.")
                return

            ws = wb[hoja]
            print("\nLista de productos registrados:\n")
            print(f"{'ID':<10} {'Nombre':<25} {'Cantidad':<10} {'Precio':<10}")
            print("-" * 60)

            # Recorre todas las filas del archivo y muestra los productos
            for fila in range(2, ws.max_row + 1):
                id_ = ws.cell(row=fila, column=1).value
                nombre = ws.cell(row=fila, column=2).value
                cantidad = ws.cell(row=fila, column=3).value
                precio = ws.cell(row=fila, column=4).value
                print(f"{str(id_):<10} {str(nombre):<25} {int(cantidad):<10} ${float(precio):<10.2f}")
        except Exception as e:
            print(f"Error al mostrar productos: {type(e).__name__} - {e}")

    def cargar_desde_excel(self, nombre_archivo, hoja):
        # Crea una instancia de la clase Archivo para manejar la carga
        archivo = Archivo()

        # Llama al método que carga los productos desde el archivo Excel
        archivo.cargar_desde_excel(nombre_archivo, hoja)

        # Verifica que los productos cargados sean una lista válida
        if isinstance(archivo.productos, list):
            # Asigna los productos cargados a la lista en memoria
            self.productos = archivo.productos
            print(f"{len(self.productos)} producto(s) cargados en memoria desde '{nombre_archivo}'.")
        else:
            # Si hubo un error en la carga, muestra un mensaje claro
            print("Error: los datos cargados no son válidos.")
