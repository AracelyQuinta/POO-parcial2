from openpyxl import Workbook, load_workbook  # Librería para manipular archivos Excel (.xlsx)
import os  # Para verificar existencia de archivos y manejar rutas

# Verifica si el archivo Excel y la hoja existen; si no, los crea con encabezados
def verificar_o_crear_excel(nombre_archivo, hoja):
    try:
        # Verifica si el archivo ya existe en el sistema
        if os.path.exists(nombre_archivo):
            wb = load_workbook(nombre_archivo)  # Carga el archivo existente

            # Si la hoja no existe dentro del archivo, la crea y agrega encabezados
            if hoja not in wb.sheetnames:
                ws = wb.create_sheet(title=hoja)
                ws.append(["ID", "Nombre", "Cantidad", "Precio"])  # Encabezados estándar
                wb.save(nombre_archivo)
                print(f"Hoja '{hoja}' creada en el archivo existente.")
            else:
                # Si el archivo y la hoja ya existen, se informa al usuario
                print(f"Archivo '{nombre_archivo}' y hoja '{hoja}' ya existen.")
        else:
            # Si el archivo no existe, se crea uno nuevo con la hoja y encabezados
            wb = Workbook()
            ws = wb.active  # Obtiene la hoja activa por defecto
            ws.title = hoja  # Renombra la hoja activa
            ws.append(["ID", "Nombre", "Cantidad", "Precio"])  # Encabezados estándar
            wb.save(nombre_archivo)
            print(f"Archivo '{nombre_archivo}' creado con la hoja '{hoja}'.")

    # Manejo de error si el archivo está protegido o no se tienen permisos
    except PermissionError:
        print(f"No tienes permisos para crear o modificar '{nombre_archivo}'.")
