import os
from openpyxl import Workbook, load_workbook
from producto import Producto

class Archivo:
    def __init__(self, productos=None):
        self.productos =self.productos = productos if productos else []

    def guardar_en_excel_xlsx(self, nombre_archivo):
        try:
            if os.path.exists(nombre_archivo):
                wb = load_workbook(nombre_archivo)
                ws = wb.active
            else:
                wb = Workbook()
                ws = wb.active
                ws.title = "Inventario"
                ws.append(["ID", "Nombre", "Cantidad", "Precio"])  # Encabezado

            for producto in self.productos:
                ws.append([
                    producto.getid(),
                    producto.getnombre(),
                    producto.getcantidad(),
                    producto.getprecio()
                ])

            wb.save(nombre_archivo)
            print(f"Se ha guardado exitosamente en '{nombre_archivo}' como archivo Excel.")

        except PermissionError:
            print(f"Error: No tienes permisos para modificar '{nombre_archivo}'.")
        except Exception as e:
            print(f" Error inesperado: {type(e).__name__} - {e}")

    def cargar_desde_excel(self, nombre_archivo, hoja):
        if os.path.exists(nombre_archivo):
            wb = load_workbook(nombre_archivo)
            ws = wb[hoja]

            for fila in range(2, ws.max_row + 1):  # Ignora encabezado
                id_ = str(ws.cell(row=fila, column=1).value)
                nombre = str(ws.cell(row=fila, column=2).value)
                cantidad = int(ws.cell(row=fila, column=3).value)
                precio = float(ws.cell(row=fila, column=4).value)

                productos_leidos = Producto(id_, nombre, cantidad, precio)
                self.productos.append(productos_leidos)

    def leer_archivo(ruta):
        try:
            with open(ruta, 'r', encoding='utf-8') as archivo:
                contenido = archivo.read()
                print("Archivo leído correctamente.")
                return contenido
        except FileNotFoundError:
            print(f" Error: El archivo '{ruta}' no existe.")
        except PermissionError:
            print(f" Error: No tienes permisos para acceder al archivo '{ruta}'.")
        except Exception as e:
            print(f" Error inesperado: {type(e).__name__} - {e}")


    def escribir_archivo(ruta, datos):
        try:
            with open(ruta, 'w', encoding='utf-8') as archivo:
                archivo.write(datos)
                print(" Datos escritos correctamente.")
        except PermissionError:
            print(f" Error: No tienes permisos para escribir en '{ruta}'.")
        except Exception as e:
            print(f" Error inesperado: {type(e).__name__} - {e}")


