from openpyxl import Workbook, load_workbook
import os

def verificar_o_crear_excel(nombre_archivo, hoja):
    try:
        if os.path.exists(nombre_archivo):
            wb = load_workbook(nombre_archivo)
            if hoja not in wb.sheetnames:
                ws = wb.create_sheet(title=hoja)
                ws.append(["ID", "Nombre", "Cantidad", "Precio"])
                wb.save(nombre_archivo)
                print(f"Hoja '{hoja}' creada en el archivo existente.")
            else:
                print(f"Archivo '{nombre_archivo}' y hoja '{hoja}' ya existen.")
        else:
            wb = Workbook()
            ws = wb.active
            ws.title = hoja
            ws.append(["ID", "Nombre", "Cantidad", "Precio"])
            wb.save(nombre_archivo)
            print(f"Archivo '{nombre_archivo}' creado, tiene el nombre de'{hoja}'.")
    except PermissionError:
        print(f"No tienes permisos para crear o modificar '{nombre_archivo}'.")



