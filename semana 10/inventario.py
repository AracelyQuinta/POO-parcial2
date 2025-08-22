'''
Clase Inventario:
Atributos: Una lista de productos.
Métodos:
Añadir nuevo producto (asegurarse de que el ID sea único).
Eliminar producto por ID.
Actualizar cantidad o precio de un producto por ID.
Buscar producto(s) por nombre (puede haber nombres similares).
Mostrar todos los productos en el inventario.
'''
from openpyxl import Workbook, load_workbook
from producto import Producto  # Importa la clase Producto definida previamente
from manipulacion_de_archivo import Archivo

class Inventario:
    def __init__(self):
        self.productos = []  # Lista en memoria para almacenar productos temporalmente

    def añadir_producto(self, producto=None):  # Permite añadir un producto desde consola
        nuevo_id = input("Ingrese el ID: ").strip()

        # Verifica si el ID ya existe en la lista de productos
        id_duplicado = False
        for n in self.productos:
            if (n.getid()) == nuevo_id:
                id_duplicado = True
                break

        if id_duplicado:
            print(f"Ya existe un producto con el ID {nuevo_id}.")
        else:
            # Solicita los datos del nuevo producto
            nombre = input("Nombre del producto: ")
            cantidad = int(input("Cantidad del producto: "))
            precio = float(input("Precio del producto: "))
            nuevo_producto = Producto(nuevo_id, nombre, cantidad, precio)
            self.productos.append(nuevo_producto)  # Lo añade a la lista en memoria
            print(f"Producto '{nombre}' añadido correctamente.")

    def guardar_productos(self, nombre_archivo):
        archivo = Archivo(self.productos)
        archivo.guardar_en_excel_xlsx(nombre_archivo)

    def mostrar_productos_guardados(self):
        # Muestra los productos que están en memoria
        if self.productos:
            for n in self.productos:
                print(n)
        else:
            print("No hay productos guardados")

    def eliminar_producto_en_excel(self, nombre_archivo, hoja, id_a_eliminar):
        wb = load_workbook(nombre_archivo)
        ws = wb[hoja]

        fila_a_eliminar = None

        # Busca la fila que contiene el ID a eliminar
        for fila in range(2, ws.max_row + 1):  # Ignora encabezado
            id_en_excel = str(ws.cell(row=fila, column=1).value)
            if id_en_excel == id_a_eliminar:
                fila_a_eliminar = fila
                break
        try:

            if fila_a_eliminar:
                ws.delete_rows(fila_a_eliminar)  # Elimina la fila
                wb.save(nombre_archivo)
                print(f" Producto con ID {id_a_eliminar} eliminado del archivo Excel.")
            else:
                print(f" No se encontró el ID {id_a_eliminar} en el archivo Excel.")
            self.guardar_productos(nombre_archivo)
        except:
            print("No es posible eliminar el producto de Excel")

    def actualizar_producto_por_id(self, nombre_archivo, hoja, id_a_actualizar):
        wb = load_workbook(nombre_archivo)
        ws = wb[hoja]

        fila_a_actualizar = None

        # Busca la fila que contiene el ID a actualizar
        for fila in range(2, ws.max_row + 1):
            id_en_excel = str(ws.cell(row=fila, column=1).value)
            if id_en_excel == id_a_actualizar:
                fila_a_actualizar = fila
                break

        if fila_a_actualizar is not None:
            print(f"\nProducto con ID {id_a_actualizar} encontrado en la fila {fila_a_actualizar}.")

            # Muestra los datos actuales del producto
            nombre_actual = ws.cell(row=fila_a_actualizar, column=2).value
            cantidad_actual = ws.cell(row=fila_a_actualizar, column=3).value
            precio_actual = ws.cell(row=fila_a_actualizar, column=4).value

            print(f"Nombre actual: {nombre_actual}")
            print(f"Cantidad actual: {cantidad_actual}")
            print(f"Precio actual: {precio_actual}")

            # Opciones de actualización
            print("\n¿Qué desea actualizar?")
            print("1. Nombre")
            print("2. Cantidad")
            print("3. Precio")
            print("4. Todo")
            print("5. Cancelar")

            entrada = input("Ingrese las opciones separadas por coma (ej. 1,3 o 4): ").strip()
            opciones = [op.strip() for op in entrada.split(",")]

            if "5" in opciones:
                print("Cambios cancelados. No se guardó nada.")
                return

            if "4" in opciones:
                opciones = ["1", "2", "3"]  # Actualiza todo

            # Actualiza nombre si se seleccionó
            if "1" in opciones:
                nuevo_nombre = input("Ingrese el nuevo nombre: ").strip()
                if nuevo_nombre:
                    ws.cell(row=fila_a_actualizar, column=2).value = nuevo_nombre
                    print("Nombre actualizado.")
                else:
                    print("Nombre vacío. No se actualizó.")

            # Actualiza cantidad si se seleccionó
            if "2" in opciones:
                nueva_cantidad = input("Ingrese la nueva cantidad: ").strip()
                if nueva_cantidad.isdigit():
                    ws.cell(row=fila_a_actualizar, column=3).value = int(nueva_cantidad)
                    print("Cantidad actualizada.")
                else:
                    print("Cantidad inválida. No se actualizó.")

            # Actualiza precio si se seleccionó
            if "3" in opciones:
                nuevo_precio = input("Ingrese el nuevo precio: ").strip()
                if nuevo_precio.replace(".", "", 1).isdigit():
                    ws.cell(row=fila_a_actualizar, column=4).value = float(nuevo_precio)
                    print("Precio actualizado.")
                else:
                    print("Precio inválido. No se actualizó.")

            wb.save(nombre_archivo)
            print("Cambios guardados correctamente.")
        else:
            print(f"No se encontró el ID {id_a_actualizar}.")
        self.guardar_productos(nombre_archivo)

    def buscar_productos_por_nombre(self, nombre_archivo, hoja, texto_busqueda):
        wb = load_workbook(nombre_archivo)
        ws = wb[hoja]

        texto_busqueda = texto_busqueda.lower().strip()
        productos_encontrados = []

        # Busca coincidencias parciales en el nombre
        for fila in range(2, ws.max_row + 1):
            nombre = ws.cell(row=fila, column=2).value
            if nombre and texto_busqueda in nombre.lower():
                producto = {
                    "ID": ws.cell(row=fila, column=1).value,
                    "Nombre": nombre,
                    "Cantidad": ws.cell(row=fila, column=3).value,
                    "Precio": ws.cell(row=fila, column=4).value
                }
                productos_encontrados.append(producto)

        # Muestra los resultados encontrados
        if productos_encontrados:
            print(f"\nSe encontraron {len(productos_encontrados)} producto(s) que coinciden con '{texto_busqueda}':\n")
            for p in productos_encontrados:
                print(f"|ID: {p['ID']} | Nombre: {p['Nombre']} | Cantidad: {p['Cantidad']} | Precio: {p['Precio']}|")
        else:
            print(f"\nNo se encontraron productos que coincidan con '{texto_busqueda}'.")

    def mostrar_todos_los_productos(self, nombre_archivo, hoja):
        wb = load_workbook(nombre_archivo)
        ws = wb[hoja]

        print("\nLista de productos registrados:\n")
        encabezado = ["ID", "Nombre", "Cantidad", "Precio"]
        print(f"{encabezado[0]:<10} {encabezado[1]:<25} {encabezado[2]:<10} {encabezado[3]:<10}")
        print("-" * 60)

        # Recorre todas las filas del archivo y muestra los productos
        for fila in range(2, ws.max_row + 1):  # Ignora encabezado
            id_ = ws.cell(row=fila, column=1).value
            nombre = ws.cell(row=fila, column=2).value
            cantidad = ws.cell(row=fila, column=3).value
            precio = ws.cell(row=fila, column=4).value

            print(f"{str(id_):<10} {str(nombre):<25} {int(cantidad):<10} ${float(precio):<10.2f}")

    def cargar_desde_excel(self, nombre_archivo, hoja):
        archivo = Archivo()
        archivo.cargar_desde_excel(nombre_archivo, hoja)
        self.productos = archivo.productos